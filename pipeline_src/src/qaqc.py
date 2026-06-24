# =============================================================================
# qaqc.py — Rapport de Contrôle Qualité et Estimations Primaires
# =============================================================================
# Ce module produit un fichier Excel multi-onglets contenant :
#   1. Résumé exécutif (dimensions, taux de complétion global)
#   2. Log des anomalies détectées par le loader
#   3. Taux de manquants par variable analytique
#   4. Estimations primaires par bloc thématique
#   5. Note documentaire sur les variables absentes
# =============================================================================

import logging
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config import VARS_ABSENTES

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Palette couleurs
# ---------------------------------------------------------------------------
BLUE_DARK  = "1F3864"
BLUE_MED   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
ORANGE     = "C55A11"
GREEN      = "375623"
YELLOW_WARN= "FFD966"
RED_ERR    = "FF0000"
GREY_HEAD  = "D9D9D9"

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _style_header_row(ws, row_idx, n_cols, fill_hex, font_color="FFFFFF"):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = _header_fill(fill_hex)
        cell.font = _font(bold=True, color=font_color)
        cell.alignment = _center()
        cell.border = _border()

def _style_data_row(ws, row_idx, n_cols, fill_hex=None):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill_hex:
            cell.fill = _header_fill(fill_hex)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _border()

def _write_df_to_ws(ws, df, header_fill=BLUE_MED, start_row=1):
    """Écrit un DataFrame dans un worksheet avec mise en forme."""
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = _border()
            if r_idx == start_row:
                cell.fill = _header_fill(header_fill)
                cell.font = _font(bold=True, color="FFFFFF")
                cell.alignment = _center()
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # Alternance de couleurs
                if (r_idx - start_row) % 2 == 0:
                    cell.fill = _header_fill("EBF3FB")
    # Ajuster largeurs colonnes
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
    return ws


# ---------------------------------------------------------------------------
# Onglet 1 — Résumé exécutif
# ---------------------------------------------------------------------------

def _sheet_resume(wb, df_raw, df_final, qc_log):
    ws = wb.create_sheet("01_Résumé")
    ws.sheet_view.showGridLines = False

    titre_data = [
        ["RAPPORT QAQC — ENQUÊTE DÉCHETS MÉNAGERS"],
        [""],
        ["DIMENSIONS DU FICHIER BRUT"],
        ["Nombre de ménages enquêtés", df_raw.shape[0]],
        ["Nombre de variables brutes", df_raw.shape[1]],
        [""],
        ["DIMENSIONS DU FICHIER TRAITÉ"],
        ["Nombre de ménages (après nettoyage)", df_final.shape[0]],
        ["Nombre de variables analytiques construites", df_final.shape[1]],
        [""],
        ["QUALITÉ GLOBALE"],
        ["Nombre d'anomalies détectées", len(qc_log)],
        ["Taux de complétion global (%)",
         round(100 * (1 - df_final.isnull().mean().mean()), 1)],
        [""],
        ["NOTE : Variables non collectées directement"],
    ]

    for r_idx, row in enumerate(titre_data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _font(bold=True, color="FFFFFF", size=14)
                cell.fill = _header_fill(BLUE_DARK)
                cell.alignment = _center()
            elif str(val).isupper() and len(str(val)) > 5:
                cell.font = _font(bold=True, color=BLUE_DARK)
            else:
                cell.alignment = Alignment(vertical="center")

    ws.merge_cells("A1:B1")
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.row_dimensions[1].height = 30

    # Variables absentes
    start = len(titre_data) + 1
    for i, (var, note) in enumerate(VARS_ABSENTES.items(), start):
        ws.cell(row=i, column=1, value=var).font = _font(bold=True, color=ORANGE)
        ws.cell(row=i, column=2, value=note)

    return ws


# ---------------------------------------------------------------------------
# Onglet 2 — Log des anomalies
# ---------------------------------------------------------------------------

def _sheet_anomalies(wb, qc_log):
    ws = wb.create_sheet("02_Anomalies_QC")
    if not qc_log:
        ws.cell(1, 1, "Aucune anomalie détectée.")
        return ws

    df_log = pd.DataFrame(qc_log)
    df_log.columns = ["Type de contrôle", "Description", "N observations affectées", "Action appliquée"]
    _write_df_to_ws(ws, df_log, header_fill=ORANGE)

    # Colorier les lignes critiques (outlier ou erreur)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if "outlier" in str(row[0].value).lower():
            for cell in row:
                cell.fill = _header_fill("FFE4C4")

    return ws


# ---------------------------------------------------------------------------
# Onglet 3 — Taux de manquants
# ---------------------------------------------------------------------------

def _sheet_manquants(wb, df_final):
    ws = wb.create_sheet("03_Manquants")

    miss = (df_final.isnull().sum() / len(df_final) * 100).round(1).reset_index()
    miss.columns = ["Variable", "% Manquant"]
    miss["N manquants"] = df_final.isnull().sum().values
    miss["Statut"] = miss["% Manquant"].apply(
        lambda x: "⚠ Élevé (>50%)" if x > 50 else ("OK" if x < 10 else "Modéré")
    )
    miss = miss.sort_values("% Manquant", ascending=False)

    _write_df_to_ws(ws, miss, header_fill=BLUE_MED)

    # Colorier les lignes à fort taux
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        pct = row[1].value
        if isinstance(pct, (int, float)):
            if pct > 50:
                for cell in row:
                    cell.fill = _header_fill("FFD7D7")
            elif pct > 20:
                for cell in row:
                    cell.fill = _header_fill(YELLOW_WARN)

    return ws


# ---------------------------------------------------------------------------
# Onglet 4 — Estimations primaires
# ---------------------------------------------------------------------------

def _pct_dist(series, label_col="Modalité", pct_col="% (non-NaN)"):
    """Distribution en % sur valeurs non-NaN."""
    counts = series.value_counts(dropna=True)
    total = counts.sum()
    df = counts.reset_index()
    df.columns = [label_col, "N"]
    df[pct_col] = (df["N"] / total * 100).round(1)
    return df


def _sheet_estimations(wb, df):
    ws = wb.create_sheet("04_Estimations_Primaires")
    ws.sheet_view.showGridLines = False
    row = 1

    sections = []

    # --- A. Ménage ---
    sections.append(("A. CARACTÉRISTIQUES DU MÉNAGE", BLUE_DARK, [
        ("Type d'habitat",          df["type_habitat"]),
        ("Statut d'occupation",     df["statut_occupation"]),
        ("Proxy niveau de vie",     df["proxy_niveau_vie_label"]),
        ("Taille ménage (stats)",   None),  # spécial numérique
    ]))

    # --- B. CM ---
    sections.append(("B. CARACTÉRISTIQUES DU CHEF DE MÉNAGE", BLUE_MED, [
        ("Sexe du CM",                  df["cm_sexe"]),
        ("Tranche d'âge du CM",         df["cm_tranche_age"]),
        ("Type d'enseignement",         df["cm_branche_etudes"]),
        ("Niveau d'instruction",        df["cm_niveau_etudes"]),
        ("CM alphabétisé",              df["cm_est_alphabetise"]),
        ("Répondant = CM",              df["repondant_est_cm"]),
    ]))

    # --- C. Déchets ---
    sections.append(("C. GESTION DES DÉCHETS", "375623", [
        ("Mode de stockage principal",       df["mode_stockage"]),
        ("Stockage couvert",                 df["stockage_couvert"]),
        ("Tri avant évacuation",             df["tri_avant_evacuation"]),
        ("Accès benne tasseuse",             df["acces_benne_tasseuse"]),
        ("Service alternatif utilisé",       df["a_service_alternatif"]),
        ("Service payant",                   df["service_evacuation_payant"]),
        ("Tranche montant évacuation",       df["montant_evacuation_tranche"]),
        ("Accès dépotoire normalisé",        df["acces_depotoire_normalise"]),
        ("Satisfaction collecte benne",      df["satisfaction_benne"]),
        ("Satisfaction bacs quartier",       df["satisfaction_bacs_quartier"]),
        ("Dépôts sauvages observés",         df["depot_sauvage_observe"]),
        ("Connaissance UCG",                 df["connait_UCG"]),
        ("Au moins 1 traitement pratiqué",   df["au_moins_un_traitement"]),
    ]))

    # --- D. Conséquences ---
    sections.append(("D. CONSÉQUENCES SANITAIRES", ORANGE, [
        ("Aucune maladie déclarée",              df["aucune_maladie"].map({1: "Oui", 0: "Non"})),
        ("Maladie potentiellement liée déchets", df["maladie_potentiellement_liee_dechets"]),
        ("Nuisibles présents (liste)",           df["nuisibles_liste"].notna().map({True: "Nuisibles signalés", False: "Aucun"})),
    ]))

    for section_title, fill_hex, variables in sections:
        # En-tête de section
        ws.cell(row=row, column=1, value=section_title)
        ws.cell(row=row, column=1).font = _font(bold=True, color="FFFFFF", size=12)
        ws.cell(row=row, column=1).fill = _header_fill(fill_hex)
        ws.merge_cells(f"A{row}:E{row}")
        ws.row_dimensions[row].height = 22
        row += 1

        for var_label, series in variables:
            # En-tête variable
            ws.cell(row=row, column=1, value=var_label)
            ws.cell(row=row, column=1).font = _font(bold=True, color=BLUE_DARK)
            ws.cell(row=row, column=1).fill = _header_fill(GREY_HEAD)

            if series is None:
                # Cas spécial : statistiques numériques pour taille ménage
                taille = df["taille_menage"].dropna()
                stats = {
                    "Minimum": taille.min(),
                    "Maximum": taille.max(),
                    "Moyenne": round(taille.mean(), 1),
                    "Médiane": taille.median(),
                    "N valides": taille.count(),
                }
                row += 1
                hdrs = list(stats.keys())
                vals = list(stats.values())
                for c_idx, h in enumerate(hdrs, 2):
                    ws.cell(row=row, column=c_idx, value=h).font = _font(bold=True)
                row += 1
                for c_idx, v in enumerate(vals, 2):
                    ws.cell(row=row, column=c_idx, value=v)
                row += 2
                continue

            row += 1
            dist = _pct_dist(series)
            # Colonnes : Modalité | N | %
            ws.cell(row=row, column=2, value="Modalité").font = _font(bold=True)
            ws.cell(row=row, column=3, value="N").font = _font(bold=True)
            ws.cell(row=row, column=4, value="% (valeurs valides)").font = _font(bold=True)
            ws.cell(row=row, column=5, value="N total (incl. NaN)").font = _font(bold=True)
            n_total = len(series)
            row += 1
            for _, dist_row in dist.iterrows():
                ws.cell(row=row, column=2, value=str(dist_row["Modalité"]))
                ws.cell(row=row, column=3, value=int(dist_row["N"]))
                ws.cell(row=row, column=4, value=float(dist_row["% (non-NaN)"]))
                ws.cell(row=row, column=5, value=n_total)
                if (row % 2) == 0:
                    for c in range(2, 6):
                        ws.cell(row=row, column=c).fill = _header_fill("EBF3FB")
                row += 1
            row += 1  # espace entre variables

    # Largeurs colonnes
    ws.column_dimensions["A"].width = 40
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 28

    return ws


# ---------------------------------------------------------------------------
# Onglet 5 — Statistiques montant mensuel évacuation
# ---------------------------------------------------------------------------

def _sheet_montant(wb, df):
    ws = wb.create_sheet("05_Montant_Evacuation")

    montant = df["montant_mensuel_evacuation_fcfa"].dropna()
    payants = df[df["service_evacuation_payant"] == "Oui"]["montant_mensuel_evacuation_fcfa"]

    stats = pd.DataFrame({
        "Indicateur": [
            "N ménages total",
            "N ménages avec service payant",
            "% ménages avec service payant",
            "Montant mensuel moyen (payants, FCFA)",
            "Montant mensuel médian (payants, FCFA)",
            "Montant mensuel min (payants, FCFA)",
            "Montant mensuel max (payants, FCFA)",
            "Montant mensuel moyen (tous ménages, 0 si gratuit)",
        ],
        "Valeur": [
            len(df),
            len(payants),
            round(len(payants) / len(df) * 100, 1),
            round(payants.mean(), 0) if len(payants) > 0 else "—",
            round(payants.median(), 0) if len(payants) > 0 else "—",
            round(payants.min(), 0) if len(payants) > 0 else "—",
            round(payants.max(), 0) if len(payants) > 0 else "—",
            round(df["montant_mensuel_evacuation_fcfa"].fillna(0).mean(), 0),
        ]
    })

    _write_df_to_ws(ws, stats, header_fill=BLUE_DARK)

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 25
    return ws


# ---------------------------------------------------------------------------
# Onglet 6 — Documentation variables absentes
# ---------------------------------------------------------------------------

def _sheet_doc_absentes(wb):
    ws = wb.create_sheet("06_Variables_Absentes")
    headers = ["Variable attendue", "Statut", "Note / Recommandation"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = _header_fill(ORANGE)
        cell.font = _font(bold=True, color="FFFFFF")
        cell.border = _border()

    for r_idx, (var, note) in enumerate(VARS_ABSENTES.items(), 2):
        ws.cell(row=r_idx, column=1, value=var).border = _border()
        ws.cell(row=r_idx, column=2, value="ABSENT — Non collectée").border = _border()
        ws.cell(row=r_idx, column=2).font = _font(color=ORANGE)
        ws.cell(row=r_idx, column=3, value=note).border = _border()

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70
    return ws


# ---------------------------------------------------------------------------
# Point d'entrée du module
# ---------------------------------------------------------------------------

def run(df_raw, df_final, qc_log, output_path):
    """Génère le rapport QAQC complet en Excel."""
    wb = Workbook()
    # Supprimer la feuille vide par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _sheet_resume(wb, df_raw, df_final, qc_log)
    _sheet_anomalies(wb, qc_log)
    _sheet_manquants(wb, df_final)
    _sheet_estimations(wb, df_final)
    _sheet_montant(wb, df_final)
    _sheet_doc_absentes(wb)

    wb.save(output_path)
    logger.info(f"Rapport QAQC sauvegardé : {output_path}")
